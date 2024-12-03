import openpyxl
import pandas as pd
from tkinter import Tk, filedialog
import xlrd
from openpyxl.styles import PatternFill
import datetime
from pathlib import Path



def dosya_sec():
    root = Tk()
    root.withdraw()  # Tkinter penceresini gizle

    dosya_yolu = filedialog.askopenfilename(
        title="Lütfen bir dosya seçin",
        filetypes=[("Excel Dosyaları", "*.xls;*.xlsx"), ("Tüm Dosyalar", "*.*")]
    )

    return dosya_yolu

# Kullanıcıdan dosya seçmesini iste
dosya_yolu = dosya_sec()


if dosya_yolu.endswith('.xls'):

    # .xls dosyasını oku
    book = xlrd.open_workbook(dosya_yolu)
    sheet = book.sheet_by_index(0)  # İlk sayfayı kullanıyoruz, eğer farklı bir sayfa kullanacaksanız indeksi değiştirebilirsiniz

    # Tüm satırları ve sütunları oku
    tum_satirlar = [sheet.row_values(i) for i in range(sheet.nrows)]
    tum_sutunlar = [sheet.col_values(i) for i in range(sheet.ncols)]

    # Sonuçları DataFrame'e yükle
    df_satirlar = pd.DataFrame(tum_satirlar)
    df_sutunlar = pd.DataFrame(tum_sutunlar)

    # Yeni bir Excel dosyası oluştur
    #dosya_yolu = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")])
    
    # Dosya adını sabit olarak ayarla
    dosya_yolu = dosya_yolu.replace(dosya_yolu.split("/")[-1], "ulaş.xlsx")

    # Verileri yeni Excel dosyasına yaz
    with pd.ExcelWriter(dosya_yolu, engine='xlsxwriter') as writer:
        df_satirlar.to_excel(writer, sheet_name='Satirlar', index=False, header=None)
        df_sutunlar.to_excel(writer, sheet_name='Sutunlar', index=False, header=None)

#elif dosya_yolu.endswith('.xlsx'):




# Excel dosyasını okuma
workbook = openpyxl.load_workbook(dosya_yolu)
sheet = workbook.active
max_rowilk = sheet.max_row
max_columnilk = sheet.max_column

# Yeni Excel dosyasını oluştur
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
max_rowson = sheet.max_row
max_columnson = sheet.max_column
max_columnson = 1
bugunun_tarihi = datetime.datetime.now().strftime("%d%m")
dosya_adi = f"{bugunun_tarihi} Otto.xls"

# Masaüstü yolunu oluştur
desktop_path = Path.home() / "Desktop"
dosya_adi = desktop_path / dosya_adi


# kopyalanacak sayfanın en sonuna not bırakıyoruz
sheet.cell(row=max_rowilk+1, column=1, value="ulas tan sevgiler")
workbook.save(dosya_yolu)


for sira in range(1,16):
    for row_number in range(1, max_rowilk + 1):
        #current_value = sheet.cell(row=row_number, column=1).value

        if sira == 1 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 2 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 3 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 4 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
        

        elif sira == 5 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 6 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
        

        elif sira == 7 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
        

        elif sira == 8 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 9 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 10 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value == None and sheet.cell(row=row_number+10, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 11 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value == None and sheet.cell(row=row_number+10, column=1).value == None and sheet.cell(row=row_number+11, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onbirinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 12 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value == None and sheet.cell(row=row_number+10, column=1).value == None and sheet.cell(row=row_number+11, column=1).value == None and sheet.cell(row=row_number+12, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onbirinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 13 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value == None and sheet.cell(row=row_number+10, column=1).value == None and sheet.cell(row=row_number+11, column=1).value == None and sheet.cell(row=row_number+12, column=1).value == None and sheet.cell(row=row_number+13, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onbirinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onüçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 14 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value == None and sheet.cell(row=row_number+10, column=1).value == None and sheet.cell(row=row_number+11, column=1).value == None and sheet.cell(row=row_number+12, column=1).value == None and sheet.cell(row=row_number+13, column=1).value == None and sheet.cell(row=row_number+14, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onbirinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onüçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ondördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1


        elif sira == 15 and sheet.cell(row=row_number, column=1).value != None and sheet.cell(row=row_number+1, column=1).value == None and sheet.cell(row=row_number+2, column=1).value == None and sheet.cell(row=row_number+3, column=1).value == None and sheet.cell(row=row_number+4, column=1).value == None and sheet.cell(row=row_number+5, column=1).value == None and sheet.cell(row=row_number+6, column=1).value == None and sheet.cell(row=row_number+7, column=1).value == None and sheet.cell(row=row_number+8, column=1).value == None and sheet.cell(row=row_number+9, column=1).value == None and sheet.cell(row=row_number+10, column=1).value == None and sheet.cell(row=row_number+11, column=1).value == None and sheet.cell(row=row_number+12, column=1).value == None and sheet.cell(row=row_number+13, column=1).value == None and sheet.cell(row=row_number+14, column=1).value == None and sheet.cell(row=row_number+15, column=1).value != None:
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #üçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #beşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #altıncı kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #yedinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #sekizinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #dokuzuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onuncu kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onbirinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onikinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onüçüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #ondördüncü kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1
            #onbeşinci kayıt
            row_number = row_number+1
            for column_number in range(1, max_columnilk + 1):
                current_value = sheet.cell(row=row_number, column=column_number).value
                new_sheet.cell(row=max_columnson, column=column_number, value=current_value)
            max_columnson = max_columnson + 1





# Yeni Excel dosyasını kaydet
new_workbook.save(dosya_adi)


# F bloğunda SKU kontrolü yapacak kırmızıya boya

for satir in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row, min_col=6, max_col=6):  # Sütun F = 6
    for hucree in satir:
        # Hücrenin değerini kontrol et (M228E01 ile başlayanlar)
        if hucree.value and str(hucree.value).startswith('M228E01') or str(hucree.value).startswith('M228E02') or str(hucree.value).startswith('ML-102') or str(hucree.value).startswith('Z208T01') or str(hucree.value).startswith('Z218T01') or str(hucree.value).startswith('Z218K01') or str(hucree.value).startswith('Z219K02') or str(hucree.value).startswith('Z198M01') or str(hucree.value).startswith('Z198T04') or str(hucree.value).startswith('Z198T05') or str(hucree.value).startswith('L218M01') or str(hucree.value).startswith('M218B01') or str(hucree.value).startswith('M218B02') or str(hucree.value).startswith('M218B03') or str(hucree.value).startswith('L178K03') or str(hucree.value).startswith('L198K01') or str(hucree.value).startswith('D188K08') or str(hucree.value).startswith('L198K02') or str(hucree.value).startswith('M218C01') or str(hucree.value).startswith('M218C02') or str(hucree.value).startswith('Z212P01') or str(hucree.value).startswith('M212B04'):
            # Kırmızı doldur
            hucree.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

new_workbook.save(dosya_adi)
# ana sayfada ulaşı silip yerine bendeki düzenlenen kaydedilsin


"""
# Düzenlenmiş.xlsx dosyasını silme işlemi ulaş.xlsx dosyasını silmeye çalışıyorum
try:
    openpyxl.remove(workbook.split("/")[-1], "ulaş.xlsx")
except FileNotFoundError:
    print()
except Exception as e:
    print()

"""





print("Veriler başarıyla yazıldı.")